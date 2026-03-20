"""
excel_generator.py  —  6-sheet professional GST Excel report
Place in: converter/utils/excel_generator.py
"""
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colours ───────────────────────────────────────────────────────────────────
NAVY   = '0D2137'; DBLUE = '1565C0'; MBLUE = '1976D2'
LBLUE  = 'DDEEFF'; PALE  = 'F0F7FF'; DGRAY = '263238'
MGRAY  = '546E7A'; LGRAY = 'F5F5F5'; WHITE = 'FFFFFF'
DGRN   = '1B5E20'; MGRN  = '2E7D32'; LGRN  = 'E8F5E9'
MORG   = 'E65100'; LORG  = 'FFF3E0'; DRED  = 'B71C1C'
LRED   = 'FFEBEE'; TEAL  = '00695C'; LTEAL = 'E0F2F1'
AMBER  = 'E65100'; LAMGR = 'FFF8E1'; PURP  = '4A148C'


def _s(st='thin', c='BDBDBD'):
    return Side(style=st, color=c)


def _b(c='BDBDBD', st='thin'):
    s = _s(st, c)
    return Border(left=s, right=s, top=s, bottom=s)


def _f(c):
    return PatternFill('solid', fgColor=c)


def sc(cell, value='', bold=False, size=10, color='000000', bg=WHITE,
       align='left', wrap=False, border=True, bc='BDBDBD',
       italic=False, bs='thin', indent=0):
    cell.value     = value
    cell.font      = Font(name='Calibri', size=size, bold=bold,
                          color=color, italic=italic)
    cell.fill      = _f(bg)
    cell.alignment = Alignment(horizontal=align, vertical='center',
                               wrap_text=wrap, indent=indent)
    if border:
        cell.border = _b(bc, bs)


# ── All summary fields definition ──────────────────────────────────────────────
SUMMARY_SECTIONS = [
    ('SUPPLIER DETAILS', DBLUE, [
        ('Supplier Name / Company',        'SUPPLIER',        True),
        ('Supplier GSTIN',                 'GSTIN_S',         True),
        ('PAN Number',                     'PAN',             False),
        ('CIN Number',                     'CIN',             False),
        ('State Code',                     'STATE_CODE',      False),
        ('Service Tax No.',                'SERVICE_TAX',     False),
        ('TAN',                            'TAN',             False),
        ('FSSAI No.',                      'FSSAI',           False),
        ('LUT Bond No.',                   'LUT_BOND',        False),
        ('MSME Number',                    'MSME',            False),
    ]),
    ('BUYER / RECIPIENT DETAILS', MBLUE, [
        ('Recipient / Buyer Name',         'RECIPIENT',       True),
        ('Recipient GSTIN',                'GSTIN_R',         True),
        ('Recipient PAN',                  'RECIPIENT_PAN',   False),
        ('Place of Supply',                'PLACE',           False),
        ('Delivery Address',               'DELIVERY_ADDR',   False),
        ('Vendor Code',                    'VENDOR_CODE',     False),
    ]),
    ('INVOICE DETAILS', TEAL, [
        ('Invoice Number',                 'INV_NO',          True),
        ('Invoice Date',                   'INV_DATE',        True),
        ('IRN (Invoice Ref Number)',        'IRN',             False),
        ('Acknowledgement No.',            'ACK_NO',          False),
        ('Acknowledgement Date',           'ACK_DATE',        False),
        ('Challan No.',                    'CHALLAN_NO',      False),
        ('Challan Date',                   'CHALLAN_DATE',    False),
        ('E-Way Bill No.',                 'EWAY_BILL',       False),
        ('Transport Name',                 'TRANSPORT_NAME',  False),
        ('Transport ID',                   'TRANSPORT_ID',    False),
        ('Vehicle No.',                    'VEHICLE_NO',      False),
        ('Ship By',                        'SHIP_BY',         False),
        ('Customer Order No.',             'CUST_ORDER',      False),
        ('OA Number',                      'OA_NO',           False),
        ('OA Date',                        'OA_DATE',         False),
        ('DBA / OBA Number',               'DBA_NO',          False),
        ('DBA / OBA Date',                 'DBA_DATE',        False),
        ('Payment Terms',                  'PAYMENT_TERMS',   False),
        ('Mode of Dispatch',               'MODE_OF_DISPATCH',False),
        ('Dispatch From',                  'DISPATCH_FROM',   False),
        ('Dispatch Date',                  'DISPATCH_DATE',   False),
    ]),
    ('ITEM / GOODS DETAILS', MORG, [
        ('HSN / SAC Code',                 'HSN',             False),
        ('Item Description',               'ITEM_DESC',       False),
        ('Quantity',                       'QTY',             False),
        ('Unit of Measure',                'UOM',             False),
        ('Rate per Unit (₹)',              'RATE',            False),
        ('GST Rate %',                     'GST_RATE',        False),
    ]),
    ('TAX & AMOUNT DETAILS', MGRN, [
        ('Taxable / Assessable Value (₹)', 'TAXABLE',         True),
        ('CGST Rate %',                    'CGST_RATE',       False),
        ('CGST Amount (₹)',                'CGST',            False),
        ('SGST Rate %',                    'SGST_RATE',       False),
        ('SGST Amount (₹)',                'SGST',            False),
        ('IGST Rate %',                    'IGST_RATE',       False),
        ('IGST Amount (₹)',                'IGST',            False),
        ('Total GST (₹)',                  'TOTAL_GST',       False),
        ('Less Discount (₹)',              'DISCOUNT',        False),
        ('Net Amount Payable (₹)',         'TOTAL',           True),
        ('Amount in Words',                'TOTAL_WORDS',     False),
    ]),
    ('BANK & PAYMENT DETAILS', PURP, [
        ('Bank Name',                      'BANK_NAME',       False),
        ('Account Number',                 'ACCOUNT_NO',      False),
        ('IFSC Code',                      'IFSC',            False),
        ('Branch',                         'BRANCH',          False),
        ('Account Holder',                 'ACCOUNT_HOLDER',  False),
        ('UPI ID',                         'UPI_ID',          False),
    ]),
]


def create_gst_excel(data, output_path, original_filename=''):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _sheet_summary(wb, data, original_filename)
    _sheet_line_items(wb, data)
    _sheet_invoice_details(wb, data)
    _sheet_tax_breakdown(wb, data)
    _sheet_gst_register(wb, data)
    _sheet_raw_ocr(wb, data)
    wb.save(output_path)
    print(f'[Excel] Saved → {output_path}')
    return output_path


# ── Sheet 1: Summary ──────────────────────────────────────────────────────────
def _sheet_summary(wb, data, orig):
    ws = wb.create_sheet('Summary', 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 52
    ws.column_dimensions['C'].width = 18

    # Banner
    ws.merge_cells('A1:C1')
    c = ws['A1']
    c.value = '  GST TAX INVOICE — EXTRACTION REPORT'
    c.font  = Font(name='Calibri', size=15, bold=True, color=WHITE)
    c.fill  = _f(NAVY)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = _b(NAVY, 'medium')
    ws.row_dimensions[1].height = 42

    ws.merge_cells('A2:C2')
    c = ws['A2']
    c.value = (f'  Generated: {datetime.now().strftime("%d %b %Y  %I:%M %p")}'
               f'   |   Source: {orig}')
    c.font      = Font(name='Calibri', size=9, italic=True, color='607D8B')
    c.fill      = _f('EEF4FB')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    # Column headers
    for col, lbl in enumerate(['  GST Field', '  Extracted Value', '  Status'], 1):
        sc(ws.cell(row=4, column=col), lbl, bold=True, size=10,
           color=WHITE, bg=DGRAY, bc=DGRAY, bs='medium')
    ws.row_dimensions[4].height = 26

    row = 5
    total_flds = 0
    found_flds = 0

    for sec_name, sec_color, fields in SUMMARY_SECTIONS:
        ws.merge_cells(f'A{row}:C{row}')
        c = ws[f'A{row}']
        c.value     = f'  {sec_name}'
        c.font      = Font(name='Calibri', size=10, bold=True, color=WHITE)
        c.fill      = _f(sec_color)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 24
        row += 1

        for lbl, key, hi in fields:
            total_flds += 1
            val   = str(data.get(key, '')).strip()
            found = bool(val)
            if found:
                found_flds += 1
            bg = PALE if hi else (LGRAY if row % 2 == 0 else WHITE)

            sc(ws.cell(row=row, column=1), lbl, bold=True, size=9,
               color=DGRAY, bg=bg, bc='CFD8DC', indent=1)
            sc(ws.cell(row=row, column=2),
               val if found else 'NOT DETECTED',
               bold=found, size=9,
               color='000000' if found else '9E9E9E',
               bg=bg, bc='CFD8DC', indent=1)
            c3 = ws.cell(row=row, column=3)
            if found:
                sc(c3, '✓  Extracted', size=9, color=DGRN,
                   bg=LGRN, align='center', bc='A5D6A7')
            else:
                sc(c3, '✗  Missing', size=9, color=DRED,
                   bg=LRED, align='center', bc='EF9A9A')
            ws.row_dimensions[row].height = 20
            row += 1

    # Accuracy score
    score = int(found_flds / total_flds * 100) if total_flds else 0
    sc_c  = DGRN if score >= 80 else (AMBER if score >= 60 else DRED)
    sc_bg = LGRN if score >= 80 else (LAMGR if score >= 60 else LRED)
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    c = ws[f'A{row}']
    c.value     = f'  Extraction Accuracy Score: {found_flds}/{total_flds} fields ({score}%)'
    c.font      = Font(name='Calibri', size=11, bold=True, color=sc_c)
    c.fill      = _f(sc_bg)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ws.row_dimensions[row].height = 26
    ws.freeze_panes = 'A5'


# ── Sheet 2: Line Items ───────────────────────────────────────────────────────
def _sheet_line_items(wb, data):
    ws = wb.create_sheet('Line Items', 1)
    ws.sheet_view.showGridLines = False

    COLS = [
        ('Sl.No.', 6), ('Description', 36), ('HSN/SAC', 12),
        ('Qty', 8), ('UOM', 8), ('Rate (₹)', 13), ('Discount', 8),
        ('Taxable (₹)', 16), ('CGST%', 8), ('CGST (₹)', 13),
        ('SGST%', 8), ('SGST (₹)', 13), ('IGST%', 8),
        ('IGST (₹)', 13), ('Total (₹)', 14),
    ]
    for i, (_, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last = get_column_letter(len(COLS))

    ws.merge_cells(f'A1:{last}1')
    sc(ws['A1'], 'GST INVOICE — ALL LINE ITEMS', bold=True, size=13,
       color=WHITE, bg=NAVY, align='center', bc=NAVY)
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f'A2:{last}2')
    sc(ws['A2'],
       f'  Invoice: {data.get("INV_NO","N/A")}  |  '
       f'Date: {data.get("INV_DATE","N/A")}  |  '
       f'Supplier GSTIN: {data.get("GSTIN_S","N/A")}',
       size=9, italic=True, color='37474F', bg='EEF4FB', indent=1)
    ws.row_dimensions[2].height = 18

    for col, (lbl, _) in enumerate(COLS, 1):
        sc(ws.cell(row=3, column=col), lbl, bold=True, size=9,
           color=WHITE, bg=DGRAY, align='center', bc=DGRAY, bs='medium')
    ws.row_dimensions[3].height = 28

    # Build items list
    items = data.get('LINE_ITEMS', [])
    if not items:
        items = [{
            'SL_NO': '1',
            'DESCRIPTION': data.get('ITEM_DESC', ''),
            'HSN': data.get('HSN', ''),
            'QTY': data.get('QTY', ''),
            'UOM': data.get('UOM', ''),
            'RATE': data.get('RATE', ''),
            'DISCOUNT': data.get('DISCOUNT', ''),
            'TAXABLE_VALUE': data.get('TAXABLE', ''),
            'CGST_RATE': data.get('CGST_RATE', ''),
            'CGST_AMT': data.get('CGST', ''),
            'SGST_RATE': data.get('SGST_RATE', ''),
            'SGST_AMT': data.get('SGST', ''),
            'IGST_RATE': data.get('IGST_RATE', ''),
            'IGST_AMT': data.get('IGST', ''),
            'TOTAL': data.get('TOTAL', ''),
        }]

    CTR = {1, 3, 4, 5, 9, 11, 13}
    RGT = {6, 7, 8, 10, 12, 14, 15}

    for idx, item in enumerate(items):
        r  = 4 + idx
        bg = PALE if idx % 2 == 0 else WHITE
        vals = [
            item.get('SL_NO', str(idx + 1)),
            item.get('DESCRIPTION', ''),
            item.get('HSN', ''),
            item.get('QTY', ''),
            item.get('UOM', ''),
            item.get('RATE', ''),
            item.get('DISCOUNT', ''),
            item.get('TAXABLE_VALUE', ''),
            item.get('CGST_RATE', ''),
            item.get('CGST_AMT', ''),
            item.get('SGST_RATE', ''),
            item.get('SGST_AMT', ''),
            item.get('IGST_RATE', ''),
            item.get('IGST_AMT', ''),
            item.get('TOTAL', ''),
        ]
        for col, val in enumerate(vals, 1):
            al = 'center' if col in CTR else ('right' if col in RGT else 'left')
            sc(ws.cell(r, col), val, size=9, bg=bg, align=al,
               bc='90CAF9', wrap=(col == 2))
        ws.row_dimensions[r].height = 22

    # Totals row
    tr = 4 + len(items)
    ws.merge_cells(f'A{tr}:G{tr}')
    sc(ws[f'A{tr}'], 'TOTALS', bold=True, size=10, color=WHITE,
       bg=DGRAY, align='center', bc=DGRAY)
    for col, key in [(8, 'TAXABLE'), (10, 'CGST'), (12, 'SGST'),
                     (14, 'IGST'), (15, 'TOTAL')]:
        sc(ws.cell(tr, col), data.get(key, ''), bold=True, size=10,
           color=DBLUE, bg=LBLUE, align='right', bc=DBLUE, bs='medium')
    ws.row_dimensions[tr].height = 24
    ws.freeze_panes = 'A4'


# ── Sheet 3: Invoice Details ──────────────────────────────────────────────────
def _sheet_invoice_details(wb, data):
    ws = wb.create_sheet('Invoice Details', 2)
    ws.sheet_view.showGridLines = False

    COLS = [
        ('#', 5), ('HSN/SAC', 13), ('Description', 34), ('Qty', 7), ('UOM', 7),
        ('Rate (₹)', 13), ('Taxable (₹)', 18), ('CGST%', 9), ('CGST (₹)', 14),
        ('SGST%', 9), ('SGST (₹)', 14), ('IGST%', 9), ('IGST (₹)', 13), ('Total (₹)', 14),
    ]
    for i, (_, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last = get_column_letter(len(COLS))

    ws.merge_cells(f'A1:{last}1')
    sc(ws['A1'], 'GST TAX INVOICE — LINE ITEM DETAILS', bold=True,
       size=13, color=WHITE, bg=NAVY, align='center', bc=NAVY)
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f'A2:{last}2')
    sc(ws['A2'],
       f'  Invoice: {data.get("INV_NO","N/A")}  |  '
       f'Supplier: {data.get("SUPPLIER","N/A")}  |  '
       f'GSTIN: {data.get("GSTIN_S","N/A")}',
       size=9, italic=True, color='37474F', bg='EEF4FB', indent=1)
    ws.row_dimensions[2].height = 18

    for col, (lbl, _) in enumerate(COLS, 1):
        sc(ws.cell(row=3, column=col), lbl, bold=True, size=9,
           color=WHITE, bg=DGRAY, align='center', bc=DGRAY, bs='medium')
    ws.row_dimensions[3].height = 28

    items = data.get('LINE_ITEMS', [])
    if not items:
        items = [{
            'SL_NO': '1', 'HSN': data.get('HSN', ''),
            'DESCRIPTION': data.get('ITEM_DESC', ''),
            'QTY': data.get('QTY', ''), 'UOM': data.get('UOM', ''),
            'RATE': data.get('RATE', ''), 'TAXABLE_VALUE': data.get('TAXABLE', ''),
            'CGST_RATE': data.get('CGST_RATE', ''), 'CGST_AMT': data.get('CGST', ''),
            'SGST_RATE': data.get('SGST_RATE', ''), 'SGST_AMT': data.get('SGST', ''),
            'IGST_RATE': data.get('IGST_RATE', ''), 'IGST_AMT': data.get('IGST', ''),
            'TOTAL': data.get('TOTAL', ''),
        }]

    CTR = {1, 3, 8, 10, 12}
    RGT = {6, 7, 9, 11, 13, 14}

    for idx, item in enumerate(items):
        r  = 4 + idx
        bg = LBLUE if idx % 2 == 0 else WHITE
        vals = [
            item.get('SL_NO', str(idx + 1)), item.get('HSN', ''),
            item.get('DESCRIPTION', ''), item.get('QTY', ''), item.get('UOM', ''),
            item.get('RATE', ''), item.get('TAXABLE_VALUE', ''),
            item.get('CGST_RATE', ''), item.get('CGST_AMT', ''),
            item.get('SGST_RATE', ''), item.get('SGST_AMT', ''),
            item.get('IGST_RATE', ''), item.get('IGST_AMT', ''),
            item.get('TOTAL', ''),
        ]
        for col, val in enumerate(vals, 1):
            al = 'center' if col in CTR else ('right' if col in RGT else 'left')
            sc(ws.cell(r, col), val, size=9, bg=bg, align=al,
               bc='90CAF9', wrap=(col == 3))
        ws.row_dimensions[r].height = 22

    tr = 4 + len(items)
    ws.merge_cells(f'A{tr}:F{tr}')
    sc(ws[f'A{tr}'], 'TOTALS', bold=True, size=10, color=WHITE,
       bg=DGRAY, align='center', bc=DGRAY)
    for col, key in [(7, 'TAXABLE'), (9, 'CGST'), (11, 'SGST'),
                     (13, 'IGST'), (14, 'TOTAL')]:
        sc(ws.cell(tr, col), data.get(key, ''), bold=True, size=10,
           color=DBLUE, bg=LBLUE, align='right', bc=DBLUE, bs='medium')
    ws.row_dimensions[tr].height = 24

    fr = tr + 2
    ws.merge_cells(f'A{fr}:D{fr}')
    sc(ws[f'A{fr}'], f'Supplier GSTIN: {data.get("GSTIN_S","N/A")}',
       bold=True, size=9, color=DBLUE, bg=LBLUE)
    ws.merge_cells(f'E{fr}:H{fr}')
    sc(ws[f'E{fr}'], f'Recipient GSTIN: {data.get("GSTIN_R","N/A")}',
       bold=True, size=9, color=DBLUE, bg=LBLUE)
    ws.merge_cells(f'I{fr}:{last}{fr}')
    sc(ws[f'I{fr}'], f'Place of Supply: {data.get("PLACE","N/A")}',
       size=9, color=DBLUE, bg=LBLUE)
    ws.freeze_panes = 'A4'


# ── Sheet 4: Tax Breakdown ────────────────────────────────────────────────────
def _sheet_tax_breakdown(wb, data):
    ws = wb.create_sheet('Tax Breakdown', 3)
    ws.sheet_view.showGridLines = False
    for col, w in zip('ABCD', [32, 22, 14, 32]):
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:D1')
    sc(ws['A1'], 'GST TAX BREAKDOWN', bold=True, size=13,
       color=WHITE, bg=DGRN, align='center', bc=DGRN)
    ws.row_dimensions[1].height = 30

    for col, lbl in enumerate(['Tax Component', 'Amount (₹)', 'Rate %', 'Notes'], 1):
        sc(ws.cell(row=2, column=col), lbl, bold=True, size=9,
           color=WHITE, bg=DGRAY, align='center', bc=DGRAY, bs='medium')
    ws.row_dimensions[2].height = 24

    rows_def = [
        ('Taxable / Assessable Value', 'TAXABLE',   '-',    'Base value before tax'),
        ('CGST (Central GST)',         'CGST',      data.get('CGST_RATE', '9%'),
         'Central Govt — 50% of GST'),
        ('SGST (State GST)',           'SGST',      data.get('SGST_RATE', '9%'),
         'State Govt — 50% of GST'),
        ('IGST (Integrated GST)',      'IGST',      data.get('IGST_RATE', '—'),
         'Inter-state supply only'),
        ('Total GST Amount',           'TOTAL_GST', '—',    'CGST + SGST (or IGST)'),
        ('Less: Discount',             'DISCOUNT',  '—',    ''),
        ('Net Amount Payable',         'TOTAL',     '—',    'Final amount due'),
    ]
    for i, (lbl, key, rate, note) in enumerate(rows_def):
        r      = 3 + i
        is_net = lbl.startswith('Net')
        is_gst = lbl.startswith('Total GST')
        bg = LTEAL if is_net else ('E8F5E9' if is_gst else
             (LGRN if i % 2 == 0 else WHITE))
        bc = TEAL if is_net else (MGRN if is_gst else 'BDBDBD')

        sc(ws.cell(r, 1), lbl, bold=(is_net or is_gst),
           size=10 if is_net else 9,
           color=TEAL if is_net else (MGRN if is_gst else '000000'),
           bg=bg, bc=bc, indent=1)
        sc(ws.cell(r, 2), data.get(key, ''),
           bold=(is_net or is_gst), size=10 if is_net else 9,
           color=TEAL if is_net else '000000',
           bg=bg, align='right', bc=bc)
        sc(ws.cell(r, 3), rate, size=9, bg=bg, align='center', bc=bc)
        sc(ws.cell(r, 4), note, size=8, color=MGRAY, italic=True, bg=bg, bc=bc)
        ws.row_dimensions[r].height = 22

    ws.merge_cells('A11:D11')
    c = ws['A11']
    c.value = (f'Invoice: {data.get("INV_NO","N/A")}  |  '
               f'Date: {data.get("INV_DATE","N/A")}  |  '
               f'Supplier GSTIN: {data.get("GSTIN_S","N/A")}  |  '
               f'Recipient GSTIN: {data.get("GSTIN_R","N/A")}')
    c.font      = Font(name='Calibri', size=8, italic=True, color='607D8B')
    c.fill      = _f(LGRAY)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[11].height = 16


# ── Sheet 5: GST Register ─────────────────────────────────────────────────────
def _sheet_gst_register(wb, data):
    ws = wb.create_sheet('GST Register', 4)
    ws.sheet_view.showGridLines = False

    COLS = [
        ('Sl.', 6), ('MSN/SAC No.', 16), ('HSN/SAC', 14), ('GST%', 9),
        ('CGST%', 9), ('CGST (₹)', 13), ('SGST%', 9), ('SGST (₹)', 13),
        ('IGST%', 9), ('IGST (₹)', 13), ('Taxable', 16), ('Total (₹)', 14),
    ]
    for i, (_, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last = get_column_letter(len(COLS))

    ws.merge_cells(f'A1:{last}1')
    sc(ws['A1'], 'GST REGISTER — COMPONENT-WISE TAX DETAILS',
       bold=True, size=12, color=WHITE, bg=AMBER, align='center', bc=AMBER)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f'A2:{last}2')
    sc(ws['A2'],
       f'  Supplier GSTIN: {data.get("GSTIN_S","N/A")}   |   '
       f'Invoice: {data.get("INV_NO","N/A")}   |   '
       f'Date: {data.get("INV_DATE","N/A")}',
       size=9, italic=True, color='37474F', bg=LAMGR, indent=1)
    ws.row_dimensions[2].height = 18

    for col, (lbl, _) in enumerate(COLS, 1):
        sc(ws.cell(row=3, column=col), lbl, bold=True, size=9,
           color=WHITE, bg=DGRAY, align='center', bc=DGRAY, bs='medium')
    ws.row_dimensions[3].height = 26

    items = data.get('LINE_ITEMS', [])
    if not items:
        hsns = [h.strip() for h in data.get('HSN', '').split(',') if h.strip()] or ['']
        items = [{
            'SL_NO': str(i + 1), 'HSN': h,
            'TAXABLE_VALUE': data.get('TAXABLE', '') if i == 0 else '',
            'CGST_AMT': data.get('CGST', '') if i == 0 else '',
            'SGST_AMT': data.get('SGST', '') if i == 0 else '',
            'IGST_AMT': data.get('IGST', '') if i == 0 else '',
            'TOTAL': data.get('TOTAL', '') if i == 0 else '',
        } for i, h in enumerate(hsns)]

    CTR = {1, 4, 5, 7, 9}
    RGT = {6, 8, 10, 11, 12}

    for idx, item in enumerate(items):
        r  = 4 + idx
        bg = LAMGR if idx % 2 == 0 else WHITE
        vals = [
            item.get('SL_NO', str(idx + 1)), '',
            item.get('HSN', ''),
            data.get('GST_RATE', ''),
            data.get('CGST_RATE', ''),
            item.get('CGST_AMT', ''),
            data.get('SGST_RATE', ''),
            item.get('SGST_AMT', ''),
            data.get('IGST_RATE', ''),
            item.get('IGST_AMT', ''),
            item.get('TAXABLE_VALUE', ''),
            item.get('TOTAL', ''),
        ]
        for col, val in enumerate(vals, 1):
            al = 'center' if col in CTR else ('right' if col in RGT else 'left')
            sc(ws.cell(r, col), val, size=9, bg=bg, align=al, bc='FFCC80')
        ws.row_dimensions[r].height = 20

    tr = 4 + len(items)
    ws.merge_cells(f'A{tr}:D{tr}')
    sc(ws[f'A{tr}'], 'TAX TOTAL', bold=True, size=10, color=WHITE,
       bg=AMBER, align='center', bc=AMBER, bs='medium')
    for col, key in [(6, 'CGST'), (8, 'SGST'), (10, 'IGST'),
                     (11, 'TAXABLE'), (12, 'TOTAL')]:
        sc(ws.cell(tr, col), data.get(key, ''), bold=True, size=10,
           color=DBLUE, bg=LAMGR, align='right', bc=AMBER, bs='medium')
    ws.row_dimensions[tr].height = 24

    wr = tr + 2
    ws.merge_cells(f'A{wr}:{last}{wr}')
    c = ws[f'A{wr}']
    c.value = (f'Net Amount Payable (in words):  '
               f'{data.get("TOTAL_WORDS", "").title() or "—"}'
               f'   (₹ {data.get("TOTAL","N/A")} /-)')
    c.font      = Font(name='Calibri', size=9, italic=True, color=DGRN)
    c.fill      = _f(LGRN)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ws.row_dimensions[wr].height = 20
    ws.freeze_panes = 'A4'


# ── Sheet 6: Raw OCR ──────────────────────────────────────────────────────────
def _sheet_raw_ocr(wb, data):
    ws = wb.create_sheet('Raw OCR Text', 5)
    ws.column_dimensions['A'].width = 130

    sc(ws['A1'], 'RAW OCR TEXT — for manual verification',
       bold=True, size=11, color=WHITE, bg=MGRAY, indent=1)
    ws.row_dimensions[1].height = 24
    sc(ws['A2'],
       f'Supplier GSTIN: {data.get("GSTIN_S","N/A")}  |  '
       f'Invoice: {data.get("INV_NO","N/A")}  |  '
       f'Date: {data.get("INV_DATE","N/A")}',
       size=9, italic=True, color=MGRAY, bg=LGRAY, indent=1)
    ws.row_dimensions[2].height = 16

    raw = data.get('raw_text', 'No text extracted')
    for i, line in enumerate(raw.split('\n'), start=3):
        c = ws.cell(row=i, column=1, value=line)
        c.font      = Font(name='Courier New', size=8, color=DGRAY)
        c.fill      = _f(LGRAY if i % 2 == 0 else WHITE)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[i].height = 14
