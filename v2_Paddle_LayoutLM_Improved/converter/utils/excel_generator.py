"""
excel_generator.py  v10.0
==========================
Place in: converter/utils/excel_generator.py

Changes v10:
  - create_batch_excel(): combined register for bulk uploads
  - Only extracted fields shown (no NOT DETECTED rows)
  - Rate fields shown when derived
"""
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY='0D2137'; DARK_BLUE='1565C0'; MID_BLUE='1976D2'; LIGHT_BLUE='DDEEFF'
PALE_BLUE='F0F7FF'; DARK_GRAY='263238'; MID_GRAY='546E7A'; LIGHT_GRAY='F5F5F5'
WHITE='FFFFFF'; DARK_GREEN='1B5E20'; MID_GREEN='2E7D32'; LIGHT_GREEN='E8F5E9'
MID_ORANGE='E65100'; LIGHT_ORANGE='FFF3E0'; DARK_RED='B71C1C'; LIGHT_RED='FFEBEE'
TEAL='00695C'; LIGHT_TEAL='E0F2F1'; AMBER='E65100'; LIGHT_AMBER='FFF8E1'
GOLD='F9A825'; LIGHT_GOLD='FFFDE7'; PURPLE='4A148C'


def _s(style='thin', color='BDBDBD'):
    return Side(style=style, color=color)

def _b(color='BDBDBD', style='thin'):
    s = _s(style, color)
    return Border(left=s, right=s, top=s, bottom=s)

def _f(color):
    return PatternFill('solid', fgColor=color)

def sc(cell, value='', bold=False, size=10, color='000000', bg=WHITE,
       align='left', wrap=False, border=True, bc='BDBDBD', italic=False, bs='thin', indent=0):
    cell.value     = value
    cell.font      = Font(name='Calibri', size=size, bold=bold, color=color, italic=italic)
    cell.fill      = _f(bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap, indent=indent)
    if border: cell.border = _b(bc, bs)


def _has_value(data, keys):
    return any(str(data.get(k,'')).strip() for k in keys)


def create_gst_excel(extracted_data, output_path, original_filename=''):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _sheet_summary(wb, extracted_data, original_filename)
    _sheet_invoice_details(wb, extracted_data)
    _sheet_tax_breakdown(wb, extracted_data)
    _sheet_gst_register(wb, extracted_data)
    _sheet_raw_ocr(wb, extracted_data)
    wb.save(output_path)
    print(f'[Excel] Saved → {output_path}')
    return output_path


def create_batch_excel(batch_data, output_path):
    """
    Create a combined Excel workbook for batch processing.
    batch_data: list of dicts with keys: filename, data (extracted_data dict)
    Sheet 1: Summary Register (one row per invoice)
    Sheets 2+: Individual invoice tabs
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _sheet_batch_register(wb, batch_data)
    for item in batch_data:
        tab_name = os.path.splitext(item['filename'])[0][:28]
        _sheet_invoice_tab(wb, item['data'], tab_name, item['filename'])
    wb.save(output_path)
    print(f'[Batch Excel] Saved {len(batch_data)} invoices → {output_path}')
    return output_path


# ═══════════════════════════════════════════════════════════════════════════
#  BATCH REGISTER SHEET
# ═══════════════════════════════════════════════════════════════════════════
def _sheet_batch_register(wb, batch_data):
    ws = wb.create_sheet('Batch Register', 0)
    ws.sheet_view.showGridLines = False

    COLS = [
        ('#',4),('File',20),('Supplier',28),('Supplier GSTIN',18),('Buyer',22),
        ('Buyer GSTIN',18),('Inv No',16),('Inv Date',14),('Place',14),
        ('Taxable (₹)',14),('CGST (₹)',12),('SGST (₹)',12),('IGST (₹)',12),
        ('Total GST (₹)',14),('Total (₹)',14),('Bank',14),('IFSC',14),
        ('HSN',10),('QTY',7),('UOM',7),('Rate (₹)',12),
    ]
    for i,(_,w) in enumerate(COLS,1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last = get_column_letter(len(COLS))

    ws.merge_cells(f'A1:{last}1')
    c = ws['A1']
    c.value     = f'  GST BATCH EXTRACTION REGISTER — {len(batch_data)} Invoices — {datetime.now().strftime("%d %b %Y %I:%M %p")}'
    c.font      = Font(name='Calibri', size=13, bold=True, color=WHITE)
    c.fill      = _f(NAVY)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 32

    for col,(lbl,_) in enumerate(COLS,1):
        sc(ws.cell(2,col), lbl, bold=True, size=9, color=WHITE,
           bg=DARK_GRAY, align='center', bc=DARK_GRAY, bs='medium')
    ws.row_dimensions[2].height = 26

    TOTALS = {'TAXABLE':0,'CGST':0,'SGST':0,'IGST':0,'TOTAL_GST':0,'TOTAL':0}
    for idx, item in enumerate(batch_data, 1):
        d = item['data']
        r = 2 + idx
        bg = PALE_BLUE if idx % 2 == 0 else WHITE

        def v(k): return str(d.get(k,'')).strip()

        vals = [
            str(idx), item['filename'], v('SUPPLIER'), v('GSTIN_S'), v('RECIPIENT'),
            v('GSTIN_R'), v('INV_NO'), v('INV_DATE'), v('PLACE'),
            v('TAXABLE'), v('CGST'), v('SGST'), v('IGST'),
            v('TOTAL_GST'), v('TOTAL'), v('BANK'), v('IFSC'),
            v('HSN'), v('QTY'), v('UOM'), v('RATE'),
        ]
        RGT = {10,11,12,13,14,21}
        CTR = {1,19,20}
        for col, val in enumerate(vals, 1):
            sc(ws.cell(r, col), val, size=9, bg=bg,
               align='right' if col in RGT else ('center' if col in CTR else 'left'),
               bc='CFD8DC')
        ws.row_dimensions[r].height = 18

        for k in TOTALS:
            try: TOTALS[k] += float(d.get(k,'0') or 0)
            except (ValueError, TypeError): pass

    # Totals row
    tr = 3 + len(batch_data)
    ws.merge_cells(f'A{tr}:I{tr}')
    sc(ws[f'A{tr}'], f'  TOTALS ({len(batch_data)} invoices)',
       bold=True, size=10, color=WHITE, bg=DARK_GRAY, align='left', bc=DARK_GRAY, bs='medium')
    for col, key in [(10,'TAXABLE'),(11,'CGST'),(12,'SGST'),(13,'IGST'),(14,'TOTAL_GST'),(15,'TOTAL')]:
        sc(ws.cell(tr,col), f'{TOTALS[key]:,.2f}', bold=True, size=10,
           color=DARK_BLUE, bg=LIGHT_BLUE, align='right', bc=DARK_BLUE, bs='medium')
    ws.row_dimensions[tr].height = 24
    ws.freeze_panes = 'A3'


# ═══════════════════════════════════════════════════════════════════════════
#  PER-INVOICE TAB (for batch workbook)
# ═══════════════════════════════════════════════════════════════════════════
def _sheet_invoice_tab(wb, data, tab_name, filename):
    ws = wb.create_sheet(tab_name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 52

    ws.merge_cells('A1:B1')
    c = ws['A1']
    c.value     = f'  {filename}'
    c.font      = Font(name='Calibri', size=11, bold=True, color=WHITE)
    c.fill      = _f(DARK_BLUE)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 26

    row = [2]
    def add(label, key, hi=False):
        val = str(data.get(key,'')).strip()
        if not val: return
        r = row[0]
        bg = PALE_BLUE if hi else (LIGHT_GRAY if r%2==0 else WHITE)
        sc(ws.cell(r,1), f'  {label}', size=9, color=DARK_GRAY, bg=bg, bc='CFD8DC', indent=1)
        sc(ws.cell(r,2), f'  {val}', bold=hi, size=9, bg=bg, bc='CFD8DC', indent=1)
        ws.row_dimensions[r].height = 18
        row[0] += 1

    add('Supplier', 'SUPPLIER', True); add('Supplier GSTIN','GSTIN_S',True)
    add('PAN','PAN'); add('CIN','CIN'); add('State Code','STATE_CODE')
    add('Recipient','RECIPIENT',True); add('Recipient GSTIN','GSTIN_R',True)
    add('Place of Supply','PLACE')
    add('Invoice No','INV_NO',True); add('Invoice Date','INV_DATE',True)
    add('IRN','IRN'); add('ACK No','ACK_NO'); add('ACK Date','ACK_DATE')
    add('OA No','OA_NO'); add('OA Date','OA_DATE')
    add('DBA No','DBA_NO'); add('DBA Date','DBA_DATE')
    add('Payment Terms','PAYMENT_TERMS')
    add('HSN / SAC','HSN'); add('Description','ITEM_DESC'); add('Qty','QTY')
    add('UOM','UOM'); add('Rate','RATE'); add('GST Rate %','GST_RATE')
    add('Taxable Value','TAXABLE',True); add('CGST %','CGST_RATE')
    add('CGST Amount','CGST',True); add('SGST %','SGST_RATE')
    add('SGST Amount','SGST',True); add('IGST %','IGST_RATE')
    add('IGST Amount','IGST',True); add('Total GST','TOTAL_GST',True)
    add('Net Payable','TOTAL',True); add('In Words','TOTAL_WORDS')
    add('Bank','BANK'); add('Account No','AC_NO'); add('IFSC','IFSC'); add('UPI','UPI_ID')


# ═══════════════════════════════════════════════════════════════════════════
#  SINGLE INVOICE SUMMARY SHEET
# ═══════════════════════════════════════════════════════════════════════════
def _sheet_summary(wb, data, original_filename):
    ws = wb.create_sheet('Summary', 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 58
    ws.column_dimensions['C'].width = 8

    ws.merge_cells('A1:C1')
    c = ws['A1']
    c.value     = '  GST TAX INVOICE — EXTRACTION REPORT'
    c.font      = Font(name='Calibri', size=15, bold=True, color=WHITE)
    c.fill      = _f(NAVY)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 42

    ws.merge_cells('A2:C2')
    c = ws['A2']
    c.value = (f'  Generated: {datetime.now().strftime("%d %b %Y  %I:%M %p")}'
               f'   |   Source: {original_filename}')
    c.font      = Font(name='Calibri', size=9, italic=True, color='607D8B')
    c.fill      = _f('EEF4FB')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6
    for col, label in enumerate(['  Field', '  Extracted Value', ' '], 1):
        sc(ws.cell(4,col), label, bold=True, size=10,
           color=WHITE, bg=DARK_GRAY, bc=DARK_GRAY, bs='medium')
    ws.row_dimensions[4].height = 26

    row_counter = [5]

    def sec(label, color, keys):
        if not _has_value(data, keys): return False
        r = row_counter[0]
        ws.merge_cells(f'A{r}:C{r}')
        c = ws[f'A{r}']
        c.value     = f'  {label}'
        c.font      = Font(name='Calibri', size=10, bold=True, color=WHITE)
        c.fill      = _f(color)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[r].height = 22
        row_counter[0] += 1
        return True

    def row_data(label, key, hi=False):
        val = str(data.get(key,'')).strip()
        if not val: return
        r   = row_counter[0]
        idx = r - 5
        bg  = PALE_BLUE if hi else (LIGHT_GRAY if idx%2==0 else WHITE)
        sc(ws.cell(r,1), f'  {label}', size=9, color=DARK_GRAY, bg=bg, bc='CFD8DC', indent=1)
        sc(ws.cell(r,2), f'  {val}', bold=hi, size=9, color='000000', bg=bg, bc='CFD8DC', indent=1)
        sc(ws.cell(r,3), '✓', bold=True, size=10, color=MID_GREEN,
           bg=LIGHT_GREEN, align='center', bc='A5D6A7')
        ws.row_dimensions[r].height = 20
        row_counter[0] += 1

    SUPPLIER_KEYS=['SUPPLIER','GSTIN_S','PAN','CIN','STATE_CODE','SERVICE_TAX','TAN','FSSAI','LUT_BOND','MSME_NO']
    sec('SUPPLIER DETAILS', DARK_BLUE, SUPPLIER_KEYS)
    row_data('Supplier Name / Company','SUPPLIER',True); row_data('Supplier GSTIN','GSTIN_S',True)
    row_data('PAN Number','PAN'); row_data('CIN Number','CIN'); row_data('State Code','STATE_CODE')
    row_data('Service Tax No.','SERVICE_TAX'); row_data('TAN','TAN')
    row_data('FSSAI No.','FSSAI'); row_data('LUT Bond No.','LUT_BOND'); row_data('MSME Number','MSME_NO')

    BUYER_KEYS=['RECIPIENT','GSTIN_R','RECIPIENT_PAN','PLACE','DELIVERY_ADDR','VENDOR_CODE']
    sec('BUYER / RECIPIENT DETAILS', MID_BLUE, BUYER_KEYS)
    row_data('Recipient / Buyer Name','RECIPIENT',True); row_data('Recipient GSTIN','GSTIN_R',True)
    row_data('Recipient PAN','RECIPIENT_PAN'); row_data('Place of Supply','PLACE')
    row_data('Delivery Address','DELIVERY_ADDR'); row_data('Vendor Code','VENDOR_CODE')

    INV_KEYS=['INV_NO','INV_DATE','IRN','ACK_NO','ACK_DATE','CHALLAN_NO','CHALLAN_DATE',
              'EWAY_BILL','TRANSPORT','TRANSPORT_ID','VEHICLE_NO','SHIP_BY','CUST_ORDER',
              'OA_NO','OA_DATE','DBA_NO','DBA_DATE','PAYMENT_TERMS','MODE_OF_DISPATCH',
              'DISPATCH_FROM','DISPATCH_DATE']
    sec('INVOICE DETAILS', TEAL, INV_KEYS)
    row_data('Invoice Number','INV_NO',True); row_data('Invoice Date','INV_DATE',True)
    row_data('IRN (Invoice Ref Number)','IRN'); row_data('Acknowledgement No.','ACK_NO')
    row_data('Acknowledgement Date','ACK_DATE'); row_data('Challan No.','CHALLAN_NO')
    row_data('Challan Date','CHALLAN_DATE'); row_data('E-Way Bill No.','EWAY_BILL')
    row_data('Transport Name','TRANSPORT'); row_data('Transport ID','TRANSPORT_ID')
    row_data('Vehicle No.','VEHICLE_NO'); row_data('Ship By','SHIP_BY')
    row_data('Customer Order No.','CUST_ORDER'); row_data('OA Number','OA_NO')
    row_data('OA Date','OA_DATE'); row_data('DBA / OBA Number','DBA_NO')
    row_data('DBA / OBA Date','DBA_DATE'); row_data('Payment Terms','PAYMENT_TERMS')
    row_data('Mode of Dispatch','MODE_OF_DISPATCH'); row_data('Dispatch From','DISPATCH_FROM')
    row_data('Dispatch Date','DISPATCH_DATE')

    ITEM_KEYS=['HSN','ITEM_DESC','QTY','UOM','RATE','GST_RATE']
    sec('ITEM / GOODS DETAILS', MID_ORANGE, ITEM_KEYS)
    row_data('HSN / SAC Code','HSN'); row_data('Item Description','ITEM_DESC')
    row_data('Quantity','QTY'); row_data('Unit of Measure','UOM')
    row_data('Rate per Unit (₹)','RATE'); row_data('GST Rate %','GST_RATE')

    TAX_KEYS=['TAXABLE','CGST_RATE','CGST','SGST_RATE','SGST','IGST_RATE','IGST','TOTAL_GST','DISCOUNT','TOTAL']
    sec('TAX & AMOUNT DETAILS', MID_GREEN, TAX_KEYS)
    row_data('Taxable / Assessable Value (₹)','TAXABLE',True)
    row_data('CGST Rate %','CGST_RATE'); row_data('CGST Amount (₹)','CGST',True)
    row_data('SGST Rate %','SGST_RATE'); row_data('SGST Amount (₹)','SGST',True)
    row_data('IGST Rate %','IGST_RATE'); row_data('IGST Amount (₹)','IGST',True)
    row_data('Total GST (₹)','TOTAL_GST',True); row_data('Less Discount (₹)','DISCOUNT')
    row_data('Net Amount Payable (₹)','TOTAL',True); row_data('Amount in Words','TOTAL_WORDS')

    BANK_KEYS=['BANK','AC_NO','IFSC','UPI_ID']
    sec('BANK & PAYMENT DETAILS', PURPLE, BANK_KEYS)
    row_data('Bank Name','BANK'); row_data('Account Number','AC_NO')
    row_data('IFSC Code','IFSC'); row_data('UPI ID','UPI_ID')

    all_keys = SUPPLIER_KEYS+BUYER_KEYS+INV_KEYS+ITEM_KEYS+TAX_KEYS+BANK_KEYS
    found  = sum(1 for k in all_keys if data.get(k))
    total  = len(all_keys)
    score  = int(found/total*100)
    sc_col = DARK_GREEN if score>=75 else (GOLD if score>=50 else DARK_RED)
    sc_bg  = LIGHT_GREEN if score>=75 else (LIGHT_GOLD if score>=50 else LIGHT_RED)

    row_counter[0] += 1
    r = row_counter[0]
    ws.merge_cells(f'A{r}:C{r}')
    c = ws[f'A{r}']
    c.value     = (f'  Extracted {found} of {total} possible fields  ({score}% detection rate)')
    c.font      = Font(name='Calibri', size=11, bold=True, color=sc_col)
    c.fill      = _f(sc_bg)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ws.row_dimensions[r].height = 28
    ws.freeze_panes = 'A5'


# ═══════════════════════════════════════════════════════════════════════════
#  REMAINING SHEETS (Invoice Details, Tax Breakdown, GST Register, Raw OCR)
# ═══════════════════════════════════════════════════════════════════════════
def _sheet_invoice_details(wb, data):
    ws = wb.create_sheet('Invoice Details', 1)
    ws.sheet_view.showGridLines = False
    COLS=[('#',5),('HSN/SAC',13),('Description',38),('Qty',7),('UOM',7),('Rate (₹)',13),
          ('Taxable (₹)',18),('CGST %',9),('CGST Amt (₹)',14),('SGST %',9),('SGST Amt (₹)',14),
          ('IGST %',9),('IGST Amt (₹)',13),('Total (₹)',14)]
    for i,(_,w) in enumerate(COLS,1): ws.column_dimensions[get_column_letter(i)].width=w
    last=get_column_letter(len(COLS))
    ws.merge_cells(f'A1:{last}1')
    sc(ws['A1'],'GST TAX INVOICE — LINE ITEM DETAILS',bold=True,size=13,color=WHITE,bg=NAVY,align='center',bc=NAVY)
    ws.row_dimensions[1].height=30
    ws.merge_cells(f'A2:{last}2')
    sc(ws['A2'],f'  Invoice: {data.get("INV_NO","—")}  |  Date: {data.get("INV_DATE","—")}  |  Supplier GSTIN: {data.get("GSTIN_S","—")}',
       size=9,italic=True,color='37474F',bg='EEF4FB',indent=1)
    ws.row_dimensions[2].height=18
    for col,(lbl,_) in enumerate(COLS,1):
        sc(ws.cell(3,col),lbl,bold=True,size=9,color=WHITE,bg=DARK_GRAY,align='center',bc=DARK_GRAY,bs='medium')
    ws.row_dimensions[3].height=28
    ROW=['1',data.get('HSN',''),data.get('ITEM_DESC',''),data.get('QTY',''),data.get('UOM',''),
         data.get('RATE',''),data.get('TAXABLE',''),data.get('CGST_RATE',''),data.get('CGST',''),
         data.get('SGST_RATE',''),data.get('SGST',''),data.get('IGST_RATE',''),data.get('IGST',''),data.get('TOTAL','')]
    CTR={1,4,5,8,10,12}; RGT={6,7,9,11,13,14}
    for col,val in enumerate(ROW,1):
        sc(ws.cell(4,col),val,size=9,bg=LIGHT_BLUE,
           align='center' if col in CTR else('right' if col in RGT else 'left'),bc='90CAF9')
    ws.row_dimensions[4].height=22
    ws.merge_cells('A5:F5')
    sc(ws['A5'],'TOTALS',bold=True,size=10,color=WHITE,bg=DARK_GRAY,align='center',bc=DARK_GRAY)
    for col,key in [(7,'TAXABLE'),(9,'CGST'),(11,'SGST'),(13,'IGST'),(14,'TOTAL')]:
        sc(ws.cell(5,col),data.get(key,''),bold=True,size=10,color=DARK_BLUE,bg=LIGHT_BLUE,align='right',bc=DARK_BLUE,bs='medium')
    ws.row_dimensions[5].height=24
    ws.freeze_panes='A4'


def _sheet_tax_breakdown(wb, data):
    ws = wb.create_sheet('Tax Breakdown', 2)
    ws.sheet_view.showGridLines = False
    for col,w in zip('ABCD',[32,22,14,32]): ws.column_dimensions[col].width=w
    ws.merge_cells('A1:D1')
    sc(ws['A1'],'GST TAX BREAKDOWN',bold=True,size=13,color=WHITE,bg=DARK_GREEN,align='center',bc=DARK_GREEN)
    ws.row_dimensions[1].height=30
    for col,lbl in enumerate(['Tax Component','Amount (₹)','Rate %','Notes'],1):
        sc(ws.cell(2,col),lbl,bold=True,size=9,color=WHITE,bg=DARK_GRAY,align='center',bc=DARK_GRAY,bs='medium')
    ws.row_dimensions[2].height=24
    gstin_s=data.get('GSTIN_S',''); gstin_r=data.get('GSTIN_R','')
    is_inter=(len(gstin_s)>=2 and len(gstin_r)>=2 and gstin_s[:2]!=gstin_r[:2]) or bool(data.get('IGST'))
    rows=[('Taxable / Assessable Value','TAXABLE','—','Base value before GST')]
    if is_inter:
        rows+=[('IGST (Integrated GST)','IGST',data.get('IGST_RATE','18%'),'Inter-state — full GST')]
    else:
        rows+=[('CGST (Central GST)','CGST',data.get('CGST_RATE','9%'),'Central — 50% of GST'),
               ('SGST (State GST)','SGST',data.get('SGST_RATE','9%'),'State — 50% of GST')]
    rows+=[('Total GST Amount','TOTAL_GST','—','CGST+SGST or IGST'),('Net Amount Payable','TOTAL','—','Final amount')]
    r=3
    for i,(label,key,rate,note) in enumerate(rows):
        val=data.get(key,'')
        if not val: continue
        is_net=label.startswith('Net'); is_gst=label.startswith('Total GST')
        bg=LIGHT_TEAL if is_net else('E8F5E9' if is_gst else(LIGHT_GREEN if i%2==0 else WHITE))
        bc=TEAL if is_net else(MID_GREEN if is_gst else 'BDBDBD')
        sc(ws.cell(r,1),label,bold=(is_net or is_gst),size=10 if is_net else 9,
           color=TEAL if is_net else'000000',bg=bg,bc=bc,indent=1)
        sc(ws.cell(r,2),val,bold=(is_net or is_gst),size=10 if is_net else 9,
           color=TEAL if is_net else'000000',bg=bg,align='right',bc=bc)
        sc(ws.cell(r,3),rate,size=9,bg=bg,align='center',bc=bc)
        sc(ws.cell(r,4),note,size=8,color=MID_GRAY,italic=True,bg=bg,bc=bc)
        ws.row_dimensions[r].height=22; r+=1


def _sheet_gst_register(wb, data):
    ws = wb.create_sheet('GST Register', 3)
    ws.sheet_view.showGridLines = False
    COLS=[('Sl',6),('HSN/SAC',14),('GST %',9),('CGST %',9),('CGST Amt',13),
          ('SGST %',9),('SGST Amt',13),('IGST %',9),('IGST Amt',13),('Taxable',16),('Total',14)]
    for i,(_,w) in enumerate(COLS,1): ws.column_dimensions[get_column_letter(i)].width=w
    last=get_column_letter(len(COLS))
    ws.merge_cells(f'A1:{last}1')
    sc(ws['A1'],'GST REGISTER — COMPONENT-WISE TAX DETAILS',bold=True,size=12,color=WHITE,bg=AMBER,align='center',bc=AMBER)
    ws.row_dimensions[1].height=28
    ws.merge_cells(f'A2:{last}2')
    sc(ws['A2'],f'  Supplier GSTIN: {data.get("GSTIN_S","—")}   |   Invoice: {data.get("INV_NO","—")}   |   Date: {data.get("INV_DATE","—")}',
       size=9,italic=True,color='37474F',bg=LIGHT_AMBER,indent=1)
    ws.row_dimensions[2].height=18
    for col,(lbl,_) in enumerate(COLS,1):
        sc(ws.cell(3,col),lbl,bold=True,size=9,color=WHITE,bg=DARK_GRAY,align='center',bc=DARK_GRAY,bs='medium')
    ws.row_dimensions[3].height=26
    hsn_list=[h.strip() for h in data.get('HSN','').split(',') if h.strip()] or ['']
    for idx,hsn in enumerate(hsn_list):
        r=4+idx; bg=LIGHT_AMBER if idx%2==0 else WHITE
        vals=[str(idx+1),hsn,data.get('GST_RATE',''),data.get('CGST_RATE',''),
              data.get('CGST','') if idx==0 else'',data.get('SGST_RATE',''),
              data.get('SGST','') if idx==0 else'',data.get('IGST_RATE',''),
              data.get('IGST','') if idx==0 else'',data.get('TAXABLE','') if idx==0 else'',
              data.get('TOTAL','') if idx==0 else'']
        CTR={1,3,4,6,8}; RGT={5,7,9,10,11}
        for col,val in enumerate(vals,1):
            sc(ws.cell(r,col),val,size=9,bg=bg,
               align='center' if col in CTR else('right' if col in RGT else 'left'),bc='FFCC80')
        ws.row_dimensions[r].height=20
    tr=4+len(hsn_list)
    ws.merge_cells(f'A{tr}:C{tr}')
    sc(ws[f'A{tr}'],'TAX TOTAL',bold=True,size=10,color=WHITE,bg=AMBER,align='center',bc=AMBER,bs='medium')
    for col,key in [(5,'CGST'),(7,'SGST'),(9,'IGST'),(10,'TAXABLE'),(11,'TOTAL')]:
        sc(ws.cell(tr,col),data.get(key,''),bold=True,size=10,color=DARK_BLUE,bg=LIGHT_AMBER,align='right',bc=AMBER,bs='medium')
    ws.row_dimensions[tr].height=24
    wr=tr+2
    ws.merge_cells(f'A{wr}:{last}{wr}')
    c=ws[f'A{wr}']
    c.value=f'Net Amount Payable (in words):  {data.get("TOTAL_WORDS","—")}   (₹ {data.get("TOTAL","—")} /-)'
    c.font=Font(name='Calibri',size=9,italic=True,color=DARK_GREEN)
    c.fill=_f(LIGHT_GREEN)
    c.alignment=Alignment(horizontal='left',vertical='center',indent=2)
    ws.row_dimensions[wr].height=20
    ws.freeze_panes='A4'


def _sheet_raw_ocr(wb, data):
    ws = wb.create_sheet('Raw OCR Text', 4)
    ws.column_dimensions['A'].width = 130
    sc(ws['A1'],'RAW OCR TEXT — for manual verification',bold=True,size=11,color=WHITE,bg=MID_GRAY,indent=1)
    ws.row_dimensions[1].height=24
    sc(ws['A2'],f'Supplier GSTIN: {data.get("GSTIN_S","—")}  |  Invoice: {data.get("INV_NO","—")}  |  Date: {data.get("INV_DATE","—")}',
       size=9,italic=True,color=MID_GRAY,bg=LIGHT_GRAY,indent=1)
    ws.row_dimensions[2].height=16
    raw=data.get('raw_text','No text extracted')
    for i,line in enumerate(raw.split('\n'),start=3):
        c=ws.cell(row=i,column=1,value=line)
        c.font=Font(name='Courier New',size=8,color=DARK_GRAY)
        c.fill=_f(LIGHT_GRAY if i%2==0 else WHITE)
        c.alignment=Alignment(horizontal='left',vertical='center')
        ws.row_dimensions[i].height=14
